"""
etiquetas.py — Generador de Excel para impresión masiva de etiquetas.
"""
import json
import os
import re
import pandas as pd

# ── Tipos de producto reconocidos (orden importa: más específico primero) ─────
_TIPOS = [
    "auriculares", "audifonos", "audífonos", "headset", "headphones",
    "teclado", "keyboard",
    "mouse", "ratón", "raton",
    "monitor", "pantalla", "display",
    "altavoz", "parlante", "speaker", "bocina", "bafle",
    "micrófono", "microfono", "microphone",
    "webcam", "cámara", "camara",
    "impresora", "printer",
    "tablet", "tableta",
    "cable", "adaptador", "hub",
    "batería", "bateria", "powerbank",
    "cargador", "charger",
    "disco", "ssd", "hdd", "pendrive", "memoria", "ram",
    "laptop", "portátil", "portatil", "notebook",
    "smartphone", "celular", "teléfono", "telefono",
    "joystick", "gamepad", "control",
    "soporte", "stand", "base",
    "ventilador", "cooler", "fan",
    "router", "modem",
]


def titulo_corto(nombre: str, max_len: int = 48) -> str:
    """
    Reduce el nombre del producto a ≤ max_len caracteres conservando lo esencial:
    tipo de producto + código/referencia + palabras clave adicionales.
    """
    nombre = nombre.strip()
    if len(nombre) <= max_len:
        return nombre

    nombre_lower = nombre.lower()

    # 1. Detectar tipo de producto
    tipo = ""
    for kw in _TIPOS:
        if kw in nombre_lower:
            tipo = kw.capitalize()
            break

    # 2. Detectar códigos de referencia (p.ej. BM-123, V7K100, MX-S300)
    refs = re.findall(r'\b[A-Z0-9]{2,}[-/][A-Z0-9]{2,}\b'
                      r'|\b[A-Z]{1,4}[0-9]{3,}\b'
                      r'|\b[0-9]{3,}[A-Z]{2,}\b', nombre)
    ref = refs[0] if refs else ""

    # 3. Palabras descriptivas: ignorar artículos, preposiciones y el tipo
    _STOP = {"de", "del", "con", "para", "por", "en", "la", "el", "los",
             "las", "un", "una", "y", "e", "o", "u", "a", "al"}
    palabras = [w for w in nombre.split()
                if w.lower() not in _STOP
                and tipo.lower() not in w.lower()
                and (not ref or w.upper() != ref.upper())]

    # 4. Construir candidato
    partes = [p for p in [tipo, ref] if p]
    candidato = " ".join(partes)
    for w in palabras:
        prueba = (candidato + " " + w).strip()
        if len(prueba) <= max_len:
            candidato = prueba
        else:
            break

    # Fallback: truncar con ellipsis
    if not candidato:
        candidato = nombre[:max_len - 1] + "…"

    return candidato[:max_len]

CONFIG_PATH = os.path.join(os.path.dirname(__file__), "config.json")


def _get_meta_key() -> str:
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        cfg = json.load(f)
    return cfg.get("yith_barcode_meta_key", "_ywbc_barcode_value")


def exportar_etiquetas(items: list[dict], ruta_salida: str):
    """
    Genera un archivo Excel con una fila por unidad (para impresión masiva).

    Cada elemento de `items` debe tener:
        - nombre     : str  — Nombre del producto
        - sku        : str  — SKU
        - precio     : float — Precio de venta
        - barcode    : str  — Valor del barcode YITH
        - cantidad   : int  — Número de etiquetas a generar

    Parámetro ruta_salida: path completo del .xlsx a crear.
    """
    filas = []
    for item in items:
        for _ in range(int(item["cantidad"])):
            filas.append({
                "Nombre": item["nombre"],
                "SKU": item["sku"],
                "Precio": item["precio"],
                "Barcode": item["barcode"],
            })

    if not filas:
        raise ValueError("No hay ítems para exportar.")

    df = pd.DataFrame(filas, columns=["Nombre", "SKU", "Precio", "Barcode"])
    df.to_excel(ruta_salida, index=False)
    return ruta_salida


def exportar_etiquetas_oc(items: list[dict], ruta_salida: str) -> str:
    """
    Genera etiquetas para una OC completa: una fila por unidad.

    Cada `item` debe tener:
        - nombre   : str   — Nombre completo del producto (se reduce a ≤48 chars)
        - precio   : float — Precio de venta actual (WooCommerce)
        - barcode  : str   — Valor barcode YITH
        - cantidad : int   — Unidades (genera N filas)

    Columnas del Excel: Titulo | Precio | Barcode
    """
    filas = []
    for item in items:
        titulo = titulo_corto(item["nombre"])
        for _ in range(int(item["cantidad"])):
            filas.append({
                "Titulo": titulo,
                "Precio": item["precio"],
                "Barcode": item["barcode"],
            })

    if not filas:
        raise ValueError("No hay ítems para exportar.")

    df = pd.DataFrame(filas, columns=["Titulo", "Precio", "Barcode"])
    # Ensure Barcode stays as text (openpyxl writes str columns as shared-string cells,
    # preventing Excel from auto-converting "000000038508" to the number 38508)
    df["Barcode"] = df["Barcode"].astype(str)

    df.to_excel(ruta_salida, index=False)

    # Post-process with openpyxl: mark Barcode column as Number Format "@" (Text)
    # so Excel never strips leading zeros even after the user edits the file
    from openpyxl import load_workbook
    wb = load_workbook(ruta_salida)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    if "Barcode" in headers:
        bc_col = headers.index("Barcode") + 1
        for row in ws.iter_rows(min_row=2, min_col=bc_col, max_col=bc_col):
            for cell in row:
                cell.number_format = "@"
                cell.value = str(cell.value) if cell.value is not None else ""
    wb.save(ruta_salida)

    return ruta_salida
